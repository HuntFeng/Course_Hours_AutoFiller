import datetime
import pickle
import os.path
import os
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from openpyxl import Workbook, load_workbook
import re
import json

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/calendar.readonly']

def check_credential():
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server()
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    return creds

def find_calendar_id(service):
    # calendarId = 'gkri75tpl319h0dblp04o1112s@group.calendar.google.com'
    calendars = service.calendarList().list().execute()
    for calendar in calendars['items']:
        #print(calendar['summary'], calendar['id'])
        if('@group.calendar.google.com' in calendar['id']):
            if(re.search(r'^[0-9]{6}\s', calendar['summary']) != None):
                return calendar['id']

def get_events(service, calendarId):
    # Call the Calendar API
    today = datetime.datetime.today()
    month = today.month - 1 if (today.day<=15) else today.month
    year = today.year - 1 if (month==0) else today.year
    month = 12 if (month==0) else month
    day = 31 if (month in [1,3,5,7,8, 10, 12]) else 30 
    
    start_date = datetime.datetime(year, month, 1).isoformat() + 'Z'
    end_date = datetime.datetime(year, month, day).isoformat() + 'Z'
    print('Getting events')
    events_result = service.events().list(calendarId=calendarId,
                                        timeMin=start_date,
                                        timeMax=end_date,
                                        singleEvents=True,
                                        orderBy='startTime').execute()
    events = events_result.get('items', [])
    return events

def write_to_excel(events):
    # find excel
    path = os.getcwd()
    files = os.listdir(path)
    flag = 0
    for f in files:
        if(os.path.isfile(os.path.join(path, f))):
            if('优佳教育老师工资发放表' in f):
                if(f[-5:] == '.xlsx'):
                    try:
                        wb = load_workbook(os.path.join(path, f), data_only=True)
                    except:
                        wb = load_workbook(os.path.join(path, f), data_only=True)
                    flag = 1
                    break
    if(flag == 0):
        print("cannot find the excel sheet!")
        return
        

    # load user_info.json
    with open("user_info.json", "r") as f:
        info = json.load(f)
    
    ws = wb.active
    # fill personal informations
    ws['A7'] = info['last_name']
    ws['B7'] = info['first_name']
    ws['C7'] = info['SIN']
    ws['G7'] = info['date_of_birth']
    ws['H7'] = info['address']
    ws['I7'] = info['postal_code']

    total_income = 0
    total_hours = 0
    # fill the courses hours
    n = len(events)
    for i, event in enumerate(events):
        start_date = event['start'].get('dateTime', event['end'].get('date'))
        start_date = datetime.datetime.strptime(start_date[:-6], r"%Y-%m-%dT%H:%M:%S")

        end_date = event['end'].get('dateTime', event['end'].get('date'))
        end_date = datetime.datetime.strptime(end_date[:-6], r"%Y-%m-%dT%H:%M:%S")

        row = i+13    # should be filled into this row in excel 
        for column in "BCDEFGH":
            if(column == 'B'):
                content = i   # enter the index of row in excel
            elif(column == 'C'):
                num_students = None
                if('description' in event.keys()):
                    num_students = len( event['description'].split(',') )
                    content = num_students # enter number of students
            elif(column == 'D'):
                content = event['summary']
            elif(column == 'E'):
                if('description' in event.keys()):
                    content = event['description'] # enter the student names
            elif(column == 'F'):
                # enter day-month in format like 1-June
                content = start_date.strftime("%d-%B")
            elif(column == 'G'):
                hours = (end_date - start_date).total_seconds()/3600
                content = hours
                total_hours += hours  # enter working hours
            else:
                # H column
                with open("user_info.json", "r") as f:
                    info = json.load(f)
                base_salary = info['payment']['hourly_rate'] # fixed salary 30 $/h
                increment = info['payment']['increment_per_student'] # fixed salary 30 $/h
                maximum_rate = info['payment']['maximum_rate'] # maximum rate 
                if(num_students != None):
                    salary = base_salary + increment*(num_students - 1)
                    if salary > maximum_rate:
                        salary = maximum_rate
                    content = salary*hours
                    total_income += content
    
            ws[column + str(row)] = content
    ws['B' + str(row+1)] = "总计"
    ws['H' + str(row+1)] = total_income
    ws['G' + str(row+1)] = total_hours
    ws['M7'] = total_income  
    ws['K7'] = total_hours
    wb.save("工资发放表-{0} {1}.xlsx".format(info['first_name'], info['last_name']))

def main():
    """Shows basic usage of the Google Calendar API.
    Prints the start and name of the next 10 events on the user's calendar.
    """
    creds = check_credential()
    service = build('calendar', 'v3', credentials=creds)
    calendarId = find_calendar_id(service)
    events = get_events(service, calendarId)
    write_to_excel(events)
    
    if not events:
        print('No upcoming events found.')
    for event in events:
        start = event['start'].get('dateTime', event['end'].get('date'))
        end = event['end'].get('dateTime', event['end'].get('date'))
        print(start, end, event['summary'])
if __name__ == '__main__':
    main()